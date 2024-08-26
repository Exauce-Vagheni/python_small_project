from openpyxl import load_workbook
type_chauffeur=input("Veuillez choisir le type de chauffeur à enregistrer (V) pour un voiturier et (M) pour un motard: ")
if(type_chauffeur=="V"):
    classeur=load_workbook("CHAUFFEURS_VOITURES.xlsx")
    chauffeurs=classeur["ENREGISTREMENT_CHAUFFEURS"]
    nom=input("Entrez le nom du chauffeur: ")
    postnom=input("Entrez le postnom du chauffeur: ")
    prenom=input("Entrez le prenom du chauffeur: ")
    contact=input("Entrez le contact du chauffeur: ")
    plaque=input("Entrez la plaque du chauffeur: ")
    code=len(chauffeurs["A"])
    parking=input("Entrez le parking du chauffeur: ")
    association=input("Entrez l'association du chauffeur': ")
    identification=[nom,postnom,prenom,contact,plaque,code,parking,association]
    chauffeurs.append(identification)
    classeur.save('CHAUFFEURS_VOITURES.xlsx')     
elif(type_chauffeur=="M"):
     classeur=load_workbook("MOTARDS.xlsx")
     motards=classeur["ENREGISTREMENT_MOTARDS"]
     nom=input("Entrez le nom du motard: ")
     postnom=input("Entrez le postnom du motard: ")
     prenom=input("Entrez le prenom du motard: ")
     contact=input("Entrez le contact du motard: ")
     plaque=input("Entrez la plaque du motard: ")
     code=len(motards["A"])
     parking=input("Entrez le parking du motard: ")
     association=input("Entrez l'association du motard': ")
     identification=[nom,postnom,prenom,contact,plaque,code,parking,association]
     motards.append(identification)
     classeur.save('MOTARDS.xlsx')
else:
    print("Valeur non reconnue")
    