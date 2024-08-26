from openpyxl import load_workbook
from openpyxl import Workbook
action=input("Voulez vous enregistrer(EN) ou rechercher(RC): ")
if action=="EN":
    enregistrement=input("Veuillez choisir le type d'enregistrement que vous voulez faire (MG) pour mariage, (MT) pour mort et (NS) pour naissance': ")
    
    if(enregistrement=="MG"):
        classeur1=load_workbook("NAISSANCES.xlsx")
        naissance=classeur1["LISTE_NAISSANCES"]
        colonne='B'
        naissances=tuple(cellule.value for cellule in naissance[colonne])
        nom_mari=input("Entrez les noms du mari: ")
        nom_mariee=input("Entrez les noms de la mariée: ")
        if( nom_mari in naissances and nom_mariee in naissances):
        
                classeur1=load_workbook("MARIAGES.xlsx")
                mariages=classeur1["LISTE_MARIAGES"]
                regime=input("Entrez le regime matrimonial du couple: ")
                parrain=input("Entrez les noms du parrain: ")
                code=len(mariages["A"])
                marraine=input("Entrez les noms de la marraine: ")
                commune=input("Entrez le nom de la commune ou a été celebré le mariage: ")
                religion=input("Entrez la religion des mariés: ")
                nbre_enfants=input("Entrez le nombre d'enfants hors mariage: ")
                date_mariage=input("Entrez la date du mariage: ")
                lieu=input("Entrez le lieu du mariage: ")
                attestation_info=[code,nom_mari,nom_mariee,date_mariage,regime,parrain,marraine,commune,religion,nbre_enfants,lieu]
                mariages.append(attestation_info)
                classeur1.save('MARIAGES.xlsx')
                classeur_create=Workbook()
                attestation_mariage=classeur_create.create_sheet(f"ATTESTATION {nom_mari} ET {nom_mariee}")
                entete_attestation=["CODE","NOMS DU MARI","NOMS DE LA MARIEE","DATE DE MARIAGE","REGIME MATRIMONIAL","NOMS DU PARRAIN","NOMS DE LA MARRAINE","COMMUNE","RELIGION","NOMBRE D'ENFANTS","LIEU DE MARIAGE"]
                attestation_mariage.append(entete_attestation)
                classeur_create.save(f"ACTES_MARIAGE/{nom_mari} ET {nom_mariee}.xlsx")
                classeur2=load_workbook(f"ACTES_MARIAGE/{nom_mari} ET {nom_mariee}.xlsx")
                attestation_mariage=classeur2[f"ATTESTATION {nom_mari} ET {nom_mariee}"]
                attestation_mariage.append(attestation_info)
                classeur2.save(f"ACTES_MARIAGE/{nom_mari} ET {nom_mariee}.xlsx")
        else:
            print("Toutes ces deux personnes doivent avoir un acte de naissance")
    elif(enregistrement=="MT"):
        classeur1=load_workbook("NAISSANCES.xlsx")
        naissance=classeur1["LISTE_NAISSANCES"]
        colonne='B'
        naissances=tuple(cellule.value for cellule in naissance[colonne])
        noms=input("Entrez les noms de la personne decedee: ")
        if(noms in naissances):
                classeur1=load_workbook("DECES.xlsx")
                deces=classeur1["LISTE_PERSONNES_DECEDES"]
                code=len(deces["A"])
                date_naissance=input("Entrez la date de naissance de la personne decedée: ")
                date_deces=input("Entrez la date de décès de la personne decedée: ")
                etat_civil=input("Entrez l'etat civil de la personne decedée': ")
                domicile=input("Entrez le domicile de la personne decedée': ")
                cause_deces=input("Entrez la cause de décès de la personne decedée': ")
                cimetiere=input("Entrez le cimetiere ou la personne decedée a été enterée': ")
                lieu=input("Entrez le lieu de déces: ")
                attestation_info=[code,noms,date_naissance,date_deces,etat_civil,domicile,cause_deces,cimetiere,lieu]
                deces.append(attestation_info)
                classeur1.save('DECES.xlsx')
                classeur_create=Workbook()
                attestation_deces=classeur_create.create_sheet(f"ATTESTATION DE DECES de {noms}")
                entete_attestation=["CODE","NOMS","DATE DE NAISSANCE","DATE DE DECES","ETAT CIVIL","DOMICILE","CAUSE DE DECES","CIMETIERE","LIEU DE DECES"]
                attestation_deces.append(entete_attestation)
                classeur_create.save(f"ACTES_DECES/{noms}.xlsx")
                classeur2=load_workbook(f"ACTES_DECES/{noms}.xlsx")
                attestation_deces=classeur2[f"ATTESTATION DE DECES de {noms}"]
                attestation_deces.append(attestation_info)
                classeur2.save(f"ACTES_DECES/{noms}.xlsx")
        else:
             print("Cette personne n'a pas d'acte de naissance")
    elif(enregistrement=="NS"):
        classeur1=load_workbook("NAISSANCES.xlsx")
        naissance=classeur1["LISTE_NAISSANCES"]
        code=len(naissance["A"])
        noms=input("Entrez les noms du nouveau né: ")
        date_naissance=input("Entrez la date de naissance du nouveau né: ")
        nom_pere=input("Entrez les noms du pere du nouveau né: ")
        nom_mere=input("Entrez les noms de la mère du nouveau né: ")
        domicile=input("Entrez le domicile des parents du nouveau né': ")
        hopital=input("Entrez l'hôpital dans laquelle l'enfant est né': ")
        etat=input("Entrez l'etat physique de l'enfant: ")
        lieu=input("Entrez le lieu de naissance: ")
        attestation_info=[code,noms,date_naissance,nom_pere,nom_mere,domicile,hopital,etat,lieu]
        naissance.append(attestation_info)
        classeur1.save('NAISSANCES.xlsx')
        classeur_create=Workbook()
        attestation_naissance=classeur_create.create_sheet(f"ATTESTATION DE NAISSANCE de {noms}")
        entete_attestation=["CODE","NOMS","DATE DE NAISSANCE","NOM DU PERE","NOM DE LA MERE","DOMICILE","HOPITAL","ETAT","LIEU DE NAISSANCE"]
        attestation_naissance.append(entete_attestation)
        classeur_create.save(f"ACTES_NAISSANCE/{noms}.xlsx")
        classeur2=load_workbook(f"ACTES_NAISSANCE/{noms}.xlsx")
        attestation_naissance=classeur2[f"ATTESTATION DE NAISSANCE de {noms}"]
        attestation_naissance.append(attestation_info)
        classeur2.save(f"ACTES_NAISSANCE/{noms}.xlsx")
elif action=="RC":
    classeur1=load_workbook("NAISSANCES.xlsx")
    naissance=classeur1["LISTE_NAISSANCES"]
    classeur1=load_workbook("DECES.xlsx")
    deces=classeur1["LISTE_PERSONNES_DECEDES"]
    classeur1=load_workbook("MARIAGES.xlsx")
    mariages=classeur1["LISTE_MARIAGES"]
    print("Vous ne pouvez qu'effectuer la recherche sur base des villes ou des lieux")
    ville=input("Entrez le lieu ou la ville de recherche: ")
    colonne_recherche1="I"
    valeurs_trouves=[]
    print(f"Les naissances dans la ville de {ville} sont: ")
    for ligne in naissance.iter_rows(values_only=True):
        if ville in ligne:
            print(ligne)
    print("-------------------------------------------------------------------")
    print(f"Les deces dans la ville de {ville} sont: ")
    for ligne in deces.iter_rows(values_only=True):
        if ville in ligne:
            print(ligne)
    print("-------------------------------------------------------------------")
    print(f"Les mariages dans la ville de {ville} sont: ")
    for ligne in mariages.iter_rows(values_only=True):
        if ville in ligne:
            print(ligne)
    print("-------------------------------------------------------------------")
    
    
